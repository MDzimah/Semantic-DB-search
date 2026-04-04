"""Excel data loading and preprocessing.

This module reads workbook rows, keeps the raw values intact, and builds a
search document that removes obvious repeated deciding-column text from the
free-text fields.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from hashlib import sha256
from pathlib import Path
from typing import Any, Iterable
import re
import unicodedata

from openpyxl import load_workbook


_NON_WORD_RE = re.compile(r"[^\w\s]+", re.UNICODE)
_WHITESPACE_RE = re.compile(r"\s+")


@dataclass(slots=True)
class ImportedRow:
	"""Normalized representation of one Excel row."""

	row_index: int
	values: dict[str, Any]
	display_text: str
	search_text: str
	deciding_values: dict[str, str]


@dataclass(slots=True)
class ImportedDataset:
	"""Workbook data ready for search indexing."""

	source_name: str
	sheet_name: str
	headers: list[str]
	rows: list[ImportedRow]
	fingerprint: str

	def to_dict(self) -> dict[str, Any]:
		return {
			"source_name": self.source_name,
			"sheet_name": self.sheet_name,
			"headers": self.headers,
			"rows": [asdict(row) for row in self.rows],
			"fingerprint": self.fingerprint,
		}


def normalize_text(value: Any) -> str:
	"""Return a conservative normalized text representation."""

	if value is None:
		return ""
	text = str(value).strip()
	if not text:
		return ""
	text = unicodedata.normalize("NFKD", text)
	text = "".join(character for character in text if not unicodedata.combining(character))
	text = text.lower()
	text = _NON_WORD_RE.sub(" ", text)
	text = _WHITESPACE_RE.sub(" ", text).strip()
	return text


def tokenized_variants(value: Any) -> list[str]:
	"""Generate a few normalized forms used for conservative matching."""

	normalized = normalize_text(value)
	if not normalized:
		return []

	variants = {normalized, normalized.replace(" ", "")}
	tokens = normalized.split()
	if tokens:
		variants.add(" ".join(sorted(tokens)))
	return [variant for variant in variants if variant]


def row_to_display_text(values: dict[str, Any], headers: Iterable[str]) -> str:
	pieces: list[str] = []
	for header in headers:
		value = values.get(header)
		if value is None or str(value).strip() == "":
			continue
		pieces.append(f"{header}: {value}")
	return " | ".join(pieces)


def row_to_search_text(
	values: dict[str, Any],
	headers: Iterable[str],
	deciding_columns: Iterable[str],
) -> tuple[str, dict[str, str]]:
	deciding_columns = [column for column in deciding_columns if column in values]
	deciding_values = {
		column: str(values.get(column, "")).strip()
		for column in deciding_columns
		if str(values.get(column, "")).strip()
	}

	free_text_parts: list[str] = []
	deciding_variants: list[str] = []
	for deciding_value in deciding_values.values():
		deciding_variants.extend(tokenized_variants(deciding_value))

	for header in headers:
		if header in deciding_values:
			continue
		value = values.get(header)
		if value is None:
			continue
		cleaned = normalize_text(value)
		if not cleaned:
			continue

		for deciding_variant in deciding_variants:
			if deciding_variant and deciding_variant in cleaned:
				cleaned = cleaned.replace(deciding_variant, " ")

		cleaned = _WHITESPACE_RE.sub(" ", cleaned).strip()
		if cleaned:
			free_text_parts.append(f"{header}: {cleaned}")

	search_text = " | ".join(free_text_parts).strip()
	if not search_text:
		search_text = row_to_display_text(values, headers)
		search_text = normalize_text(search_text)

	return search_text, deciding_values


def clean_row_text(
	values: dict[str, Any],
	headers: Iterable[str],
	deciding_columns: Iterable[str],
) -> tuple[str, dict[str, str]]:
	"""Build the searchable text and keep deciding values separately."""

	search_text, deciding_values = row_to_search_text(values, headers, deciding_columns)
	return search_text, deciding_values


def apply_deciding_columns(dataset: ImportedDataset, deciding_columns: Iterable[str]) -> ImportedDataset:
	"""Return a copy of the dataset with search text rebuilt for the chosen columns."""

	deciding_columns = [column for column in deciding_columns if column in dataset.headers]
	rebuilt_rows: list[ImportedRow] = []

	for row in dataset.rows:
		search_text, deciding_values = clean_row_text(row.values, dataset.headers, deciding_columns)
		rebuilt_rows.append(
			ImportedRow(
				row_index=row.row_index,
				values=row.values,
				display_text=row.display_text,
				search_text=search_text,
				deciding_values=deciding_values,
			)
		)

	fingerprint = build_dataset_fingerprint(dataset.source_name, dataset.sheet_name, dataset.headers, rebuilt_rows)
	return ImportedDataset(
		source_name=dataset.source_name,
		sheet_name=dataset.sheet_name,
		headers=dataset.headers,
		rows=rebuilt_rows,
		fingerprint=fingerprint,
	)


def load_excel_dataset(
	source: str | Path | Any,
	sheet_name: str | None = None,
	deciding_columns: Iterable[str] | None = None,
) -> ImportedDataset:
	"""Load an Excel workbook into normalized rows."""

	deciding_columns = list(deciding_columns or [])
	workbook = load_workbook(source, data_only=True, read_only=True)
	try:
		worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
		headers: list[str] = []
		rows: list[ImportedRow] = []

		for row_number, raw_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
			if row_number == 1:
				headers = [str(cell).strip() if cell is not None else f"column_{index + 1}" for index, cell in enumerate(raw_row)]
				continue

			if not any(cell is not None and str(cell).strip() for cell in raw_row):
				continue

			values = {
				headers[index]: raw_row[index]
				for index in range(min(len(headers), len(raw_row)))
			}
			display_text = row_to_display_text(values, headers)
			search_text, deciding_values = clean_row_text(values, headers, deciding_columns)
			rows.append(
				ImportedRow(
					row_index=row_number,
					values=values,
					display_text=display_text,
					search_text=search_text,
					deciding_values=deciding_values,
				)
			)

		source_name = getattr(source, "name", None) or str(source)
		fingerprint = build_dataset_fingerprint(source_name, worksheet.title, headers, rows)
		return ImportedDataset(
			source_name=source_name,
			sheet_name=worksheet.title,
			headers=headers,
			rows=rows,
			fingerprint=fingerprint,
		)
	finally:
		workbook.close()


def build_dataset_fingerprint(
	source_name: str,
	sheet_name: str,
	headers: Iterable[str],
	rows: Iterable[ImportedRow],
) -> str:
	hasher = sha256()
	hasher.update(normalize_text(source_name).encode("utf-8"))
	hasher.update(normalize_text(sheet_name).encode("utf-8"))
	for header in headers:
		hasher.update(normalize_text(header).encode("utf-8"))
	row_count = 0
	for row in rows:
		row_count += 1
		hasher.update(str(row.row_index).encode("utf-8"))
		hasher.update(row.search_text.encode("utf-8"))
	hasher.update(str(row_count).encode("utf-8"))
	return hasher.hexdigest()


def suggest_deciding_columns(headers: Iterable[str]) -> list[str]:
	"""Suggest likely deciding columns from the headers."""

	keywords = (
		"broker",
		"issuer",
		"family",
		"provider",
		"manager",
		"ticker",
		"isin",
		"fund",
		"name",
		"asset",
		"category",
		"class",
	)
	suggestions: list[str] = []
	for header in headers:
		normalized = normalize_text(header)
		if any(keyword in normalized for keyword in keywords):
			suggestions.append(header)
	return suggestions
